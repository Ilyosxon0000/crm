# import openpyxl
# import os

# # Define the path to your XLSX file (replace 'myexcel.xlsx' with the actual filename)
# excel_file = 'myexcel.xlsx'

# # Create a directory to save the extracted images
# output_directory = 'extracted_images'
# os.makedirs(output_directory, exist_ok=True)

# # Load the XLSX workbook
# workbook = openpyxl.load_workbook(excel_file, data_only=True)

# # Iterate through sheets in the XLSX file
# for sheet in workbook.sheetnames:
#     current_sheet = workbook[sheet]
#     for image in current_sheet._images:
#         image_data = image.image
#         image_filename = os.path.join(output_directory, f'image_{image.anchor.row}_{image.anchor.col}.png')

#         with open(image_filename, 'wb') as image_file:
#             image_file.write(image_data)

# print(f'Images extracted and saved in the "{output_directory}" directory.')

#Importing the modules
import openpyxl
from openpyxl_image_loader import SheetImageLoader
from PIL import ImageOps,Image


#loading the Excel File and the sheet
pxl_doc = openpyxl.load_workbook('myexcel.xls')
sheet = pxl_doc['Sheet1']

#calling the image_loader
image_loader = SheetImageLoader(sheet)

#get the image (put the cell you need instead of 'A1')
image = image_loader.get('A1')
# image = ImageOps.alpha_composite(Image.new('RGB', image.size, (255, 255, 255)), image)
image = image.convert('RGB')

# #showing the image
# image.show()

#saving the image
image.save('/home/ilyosxon/Asosiy/crm/backend/crm/extracted_images/image_name.jpg')